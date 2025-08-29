from django.shortcuts import render, get_object_or_404, redirect
from .models import Contribution
from identification.models import Membre
from django.db.models import Q, Count
from django.utils import timezone

def liste_membres_finance(request):
    membres = Membre.objects.all()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        membres = membres.filter(
            Q(nom__icontains=search_query) |
            Q(post_nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(province__icontains=search_query) |
            Q(categorie__icontains=search_query) |
            Q(statut__icontains=search_query)
        )

    # Statistiques (optionnelles)
    membres_actifs = Membre.objects.filter(statut="actif").count()
    membres_inactifs = Membre.objects.filter(statut="inactif").count()
    hommes_count = Membre.objects.filter(sexe="M").count()
    femmes_count = Membre.objects.filter(sexe="F").count()
    total_membres = membres.count()

    context = {
        'membres': membres,
        'total_membres': total_membres,
        'membres_actifs': membres_actifs,
        'membres_inactifs': membres_inactifs,
        'hommes_count': hommes_count,
        'femmes_count': femmes_count,
        'search_query': search_query,
        'show_finance_buttons': True,  # 🔑 active seulement les boutons Finance
    }

    return render(request, "liste.html", context)


##########################################################################################################
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from .models import Membre, Contribution
from .forms import ContributionForm



def historique(request, membre_id):
    membre = get_object_or_404(Membre, id=membre_id)
    contributions = membre.contributions.all().order_by('-mois')
    
    # Total des contributions
    total_contributions = contributions.aggregate(Sum('montant'))['montant__sum'] or 0
    
    # Moyenne par mois payé
    if contributions.exists():
        moyenne_mensuelle = total_contributions / contributions.count()
    else:
        moyenne_mensuelle = 0
    
    return render(request, "historique.html", {
        "membre": membre,
        "contributions": contributions,
        "total_contributions": total_contributions,
        "moyenne_mensuelle": round(moyenne_mensuelle, 2)
    })

#################################################################################################

from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from datetime import datetime
from .models import Membre, Contribution
from django.contrib import messages  # pour afficher des messages

def ajouter(request, membre_id):
    membre = get_object_or_404(Membre, id=membre_id)
    
    # Récupérer les 5 contributions les plus récentes
    contributions_recentes = Contribution.objects.filter(
        membre=membre
    ).order_by('-date_paiement')[:5]
    
    if request.method == 'POST':
        montant = request.POST.get("montant")
        mois = request.POST.get("mois")  # format YYYY-MM
        
        if montant and mois:
            try:
                # Conversion du mois en date
                mois_date = datetime.strptime(mois, "%Y-%m").date()
                
                # Vérifier si une contribution existe déjà pour ce mois
                if Contribution.objects.filter(membre=membre, mois=mois_date).exists():
                    messages.error(request, "Cette contribution pour ce mois existe déjà.")
                else:
                    Contribution.objects.create(
                        membre=membre,
                        montant=int(montant),
                        mois=mois_date
                    )
                    messages.success(request, "Contribution ajoutée avec succès.")
                    return redirect('finance:historique', membre_id=membre.id)
                
            except ValueError:
                messages.error(request, "Format de mois invalide. Utilisez YYYY-MM.")
    
    context = {
        'membre': membre,
        'contributions_recentes': contributions_recentes,
    }
    
    return render(request, 'ajouter.html', context)

#############################################################################################
def modifier_contribution(request, contribution_id):
    contribution = get_object_or_404(Contribution, id=contribution_id)
    if request.method == 'POST':
        form = ContributionForm(request.POST, instance=contribution)
        if form.is_valid():
            form.save()
            return redirect('finance:historique', membre_id=contribution.membre.id)
    else:
        form = ContributionForm(instance=contribution)
    
    return render(request, 'modifier_contribution.html', {'form': form, 'contribution': contribution})

#######################################################################################################

def supprimer_contribution(request, contribution_id):
    contribution = get_object_or_404(Contribution, id=contribution_id)
    membre_id = contribution.membre.id
    if request.method == 'POST':
        contribution.delete()
        return redirect('finance:historique', membre_id=membre_id)
    
    return render(request, 'supprimer_contribution.html', {'contribution': contribution})

#######################################################################################################
#FACTURE D'UNE SEULE COTISATION
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
import os
import qrcode
from PIL import Image
import tempfile
from django.conf import settings
from .models import Contribution, Membre
from django.contrib.staticfiles import finders

# Fonction pour convertir les nombres en lettres (version améliorée)
def nombre_en_lettres(n):
    """
    Convertit un nombre en lettres en français avec une majuscule au début.
    """
    if not isinstance(n, (int, float)) or n < 0:
        return "Nombre invalide"
    
    # Si c'est un float, on sépare les parties entière et décimale
    if isinstance(n, float):
        partie_entiere = int(n)
        partie_decimale = round((n - partie_entiere) * 100)
        if partie_decimale == 0:
            return nombre_en_lettres(partie_entiere) + " dollars américains"
        else:
            return f"{nombre_en_lettres(partie_entiere)} dollars américains et {nombre_en_lettres(partie_decimale)} cents"
    
    # Pour les entiers
    units = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
    teens = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
    tens = ["", "dix", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante-dix", "quatre-vingt", "quatre-vingt-dix"]
    big_numbers = {
        100: "cent",
        1000: "mille",
        1000000: "million",
        1000000000: "milliard"
    }

    if n == 0:
        return "Zéro"
    elif n < 10:
        return units[n].capitalize()
    elif 10 <= n < 20:
        return teens[n - 10].capitalize()
    elif 20 <= n < 100:
        if n % 10 == 0:
            return tens[n // 10].capitalize()
        elif n // 10 == 7 or n // 10 == 9:
            # Cas particuliers pour soixante-dix et quatre-vingt-dix
            base = tens[n // 10 - 1] if n // 10 == 7 else tens[n // 10 - 1]
            return f"{base.capitalize()}-{teens[n % 10 - 10]}"
        else:
            separator = "-et-" if n % 10 == 1 and n // 10 != 8 else "-"
            return f"{tens[n // 10].capitalize()}{separator}{units[n % 10]}"
    else:
        # Trouver la plus grande unité
        divisor = 1
        for d in sorted(big_numbers.keys(), reverse=True):
            if n >= d:
                divisor = d
                break
        
        quotient = n // divisor
        remainder = n % divisor
        
        if divisor == 100:
            if quotient == 1 and remainder == 0:
                return "Cent"
            elif quotient == 1:
                return f"Cent {nombre_en_lettres(remainder)}"
            elif remainder == 0:
                return f"{units[quotient].capitalize()} cents"
            else:
                return f"{units[quotient].capitalize()} cent {nombre_en_lettres(remainder)}"
        else:
            unit_name = big_numbers[divisor]
            if quotient == 1:
                prefix = "Un"
            else:
                prefix = nombre_en_lettres(quotient)
            
            if remainder == 0:
                return f"{prefix} {unit_name}"
            else:
                return f"{prefix} {unit_name} {nombre_en_lettres(remainder)}"

# Fonction pour générer un QR code
def generer_qr_code(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    img.save(temp_file.name)
    return temp_file.name

#################################################################################################################################

def generer_facture_contribution(request, pk):
    contribution = get_object_or_404(Contribution, pk=pk)
    membre = contribution.membre  

    # Préparer la réponse HTTP avec en-tête PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Facture_CONT-{contribution.pk:04d}_{membre.nom}.pdf"'

    # Créer un buffer pour le PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Définition des marges
    MARGIN_LEFT = 1.5 * cm
    MARGIN_RIGHT = 1.5 * cm
    MARGIN_TOP = 1.5 * cm
    MARGIN_BOTTOM = 1.5 * cm

    # Couleurs de la charte graphique
    primary_color = colors.HexColor("#2c3e50")  # Bleu foncé
    secondary_color = colors.HexColor("#3498db")  # Bleu
    accent_color = colors.HexColor("#e74c3c")  # Rouge
    light_color = colors.HexColor("#ecf0f1")  # Gris très clair
    dark_color = colors.HexColor("#2c3e50")  # Texte foncé

    # Styles de texte
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=primary_color,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=dark_color,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        textColor=dark_color,
        fontName='Helvetica'
    )
    
    bold_style = ParagraphStyle(
        'CustomBold',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        textColor=dark_color,
        fontName='Helvetica-Bold'
    )

    # En-tête avec deux colonnes (logo + infos société)
    p.setFillColor(light_color)
    p.rect(0, height - 3.5*cm, width, 3.5*cm, fill=True, stroke=False)
    
    # Logo (à gauche)
    try:
        logo_path = finders.find('images/logo1.png')
        if logo_path:  # si trouvé
            p.drawImage(logo_path, MARGIN_LEFT, height - 3.5*cm, width=2.5*cm, height=2.5*cm, mask='auto')
        else:
            raise FileNotFoundError("Logo non trouvé")
    except:
        # Placeholder
        p.setFillColor(secondary_color)
        p.roundRect(MARGIN_LEFT, height - 3*cm, 2.5*cm, 2.5*cm, 5, stroke=False, fill=True)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.2*cm, "LOGO")
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.6*cm, "RENEMICO")

    # Informations de la société (à droite)
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width - MARGIN_RIGHT, height - 2*cm, "REGROUPEMENT DES NEGOCIANTS MINIERS DU CONGO")
    
    p.setFont("Helvetica", 10)
    p.drawRightString(width - MARGIN_RIGHT, height - 2.7*cm, "297, Avenue Lubudi, Quartier Industriel, Manika, Kolwezi")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.1*cm, "Lualaba - RDC")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.5*cm, "Tél: +243 81 60 69 861 - Email: contact@renemico.com")

    # Position verticale initiale après l'en-tête
    y = height - 4.5*cm

    # Titre du document avec bandeau coloré
    p.setFillColor(secondary_color)
    p.rect(0, y - 0.5*cm, width, 1.2*cm, fill=True, stroke=False)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, y, "FACTURE DE COTISATION MENSUELLE")
    y -= 1.8*cm

    # Informations de la facture dans un tableau à deux colonnes
    facture_data = [
        ['N° Facture:', f"COT-{contribution.pk:04d}", 'Date:', contribution.date_paiement.strftime('%d/%m/%Y')],
        ['Mois de cotisation:', contribution.mois.strftime('%B %Y').upper(), 'Échéance:', contribution.date_paiement.strftime('%d/%m/%Y')],
    ]
    
    facture_table = Table(facture_data, colWidths=[3.5*cm, 7*cm, 3*cm, 7*cm])
    facture_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (2, 0), (2, -1), 'Helvetica-Bold', 10),
        ('BACKGROUND', (0, 0), (0, -1), light_color),
        ('BACKGROUND', (2, 0), (2, -1), light_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    facture_table.wrapOn(p, width, height)
    facture_table.drawOn(p, MARGIN_LEFT, y - 1.2*cm)
    y -= 2.5*cm

    # Informations du membre
    p.setFillColor(primary_color)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(MARGIN_LEFT, y, "INFORMATIONS DU MEMBRE")
    y -= 1*cm

    # Ligne de séparation
    p.setStrokeColor(secondary_color)
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    y -= 1.3*cm
    
    # Tableau des informations du membre
    membre_data = [
        ['Nom complet:', f"{membre.nom} {membre.post_nom} {membre.prenom}"],
        ['Adresse:', f"{membre.adresse if membre.adresse else 'Non renseignée'}"],
        ['Téléphone:', f"{membre.telephone if membre.telephone else 'Non renseigné'}"],
        ['Email:', f"{membre.email if membre.email else 'Non renseigné'}"],
    ]
    
    membre_table = Table(membre_data, colWidths=[3*cm, 12*cm])
    membre_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('BACKGROUND', (0, 0), (0, -1), light_color),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    membre_table.wrapOn(p, width, height)
    membre_table.drawOn(p, MARGIN_LEFT, y - len(membre_data)*0.7*cm)
    y -= len(membre_data)*0.9*cm + 1*cm

    # Détails de la facture
    p.setFillColor(primary_color)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(MARGIN_LEFT, y, "DÉTAILS DE LA COTISATION")
    y -= 1*cm

    # Ligne de séparation
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    y -= 1*cm
    
    # Tableau des détails
    details_data = [
        ['Description', 'Période', 'Montant (USD)'],
        ['Cotisation mensuelle', contribution.mois.strftime('%B %Y').upper(), f"{contribution.montant:.2f}"]
    ]
    
    details_table = Table(details_data, colWidths=[8*cm, 4*cm, 3.5*cm])
    details_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
    ]))
    
    details_table.wrapOn(p, width, height)
    details_table.drawOn(p, MARGIN_LEFT, y - len(details_data)*0.7*cm)
    y -= len(details_data)*0.9*cm + 1*cm

    # Montant en lettres
    # Montant en lettres
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(MARGIN_LEFT, y, "Montant en lettres:")
    p.setFont("Helvetica", 10)

    # Conversion en lettres + devise
    montant_lettres = nombre_en_lettres(contribution.montant)
    montant_lettres = f"{montant_lettres} dollars américains"

    # Ajuster la longueur du texte si nécessaire
    if len(montant_lettres) > 80:
        parts = []
        current_part = ""
        for word in montant_lettres.split():
            if len(current_part) + len(word) + 1 < 80:
                current_part += " " + word
            else:
                parts.append(current_part.strip())
                current_part = word
        parts.append(current_part.strip())
        
        for i, part in enumerate(parts):
            p.drawString(MARGIN_LEFT + 3.5*cm, y - i*0.4*cm, part)
        y -= (len(parts) - 1)*0.4*cm
    else:
        p.drawString(MARGIN_LEFT + 3.5*cm, y, montant_lettres)

    y -= 0.8*cm


    # Cadre pour le montant total
    p.setFillColor(light_color)
    p.setStrokeColor(secondary_color)
    p.setLineWidth(1.5)
    p.roundRect(width - 8*cm, y - 1.5*cm, 6.5*cm, 1.5*cm, 5, stroke=True, fill=True)
    
    p.setFillColor(primary_color)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(width - 7.5*cm, y - 0.7*cm, "TOTAL:")
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width - MARGIN_RIGHT - 0.5*cm, y - 0.7*cm, f"{contribution.montant:.2f} USD")
    y -= 2.5*cm

    # Générer et ajouter le QR Code
    qr_data = f"""
    RENEMICO - Facture de cotisation
    N°: COT-{contribution.pk:04d}
    Membre: {membre.nom} {membre.post_nom} {membre.prenom}
    Mois: {contribution.mois.strftime('%B %Y')}
    Montant: {contribution.montant} USD
    Date: {contribution.date_paiement.strftime('%d/%m/%Y')}
    """
    
    qr_code_path = generer_qr_code(qr_data)
    p.drawImage(qr_code_path, MARGIN_LEFT, y - 1.5*cm, width=3.5*cm, height=3.5*cm)
    os.unlink(qr_code_path)  # Supprimer le fichier temporaire
    
    # --- SIGNATURE RENEMICO AVEC IMAGE ---
    try:
        signature_path = finders.find("images/pca.png")  # mets ton fichier ici
        if signature_path:
            sig_width = 4.5*cm
            sig_height = 4*cm
            sig_x = MARGIN_LEFT + 6*cm   # position horizontale (ajuste selon besoin)
            sig_y = y - 1*cm             # position verticale (ajuste aussi)

            # Texte "Pour l'organisation:"
            p.setFillColor(dark_color)
            p.setFont("Helvetica", 10)
            p.drawString(MARGIN_LEFT + 4.5*cm, y - 6.5*cm, "Pour l'organisation:")

            # Ligne
            p.line(MARGIN_LEFT + 4.5*cm, y - 6.7*cm, MARGIN_LEFT + 10*cm, y - 6.7*cm)

            # Image de la signature
            p.drawImage(signature_path, sig_x, sig_y, width=sig_width, height=sig_height, mask='auto')

            # Texte "Signature et cachet"
            p.drawString(MARGIN_LEFT + 15*cm, sig_y - 0.0*cm, "Signature")

    except Exception as e:
        print("Erreur insertion signature:", e)
    
    # Notes et conditions en bas de page
    y = MARGIN_BOTTOM + 2.5*cm
    p.setStrokeColor(colors.HexColor('#bdc3c7'))
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    
    p.setFillColor(colors.HexColor('#7f8c8d'))
    p.setFont("Helvetica-Oblique", 9)
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.8*cm, "Merci pour votre confiance et votre cotisation mensuelle !")
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.3*cm, "Cette facture est générée automatiquement, pour toute question, contactez-nous à contact@renemico.com ou au +243 81 60 69 861.")

    p.showPage()
    p.save()
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


#######################################################################################################
#######################################################################################################
#HISTORIQUE DES COTISTIONS D'UN MEMBRE
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import os
import qrcode
import tempfile
from django.conf import settings
from .models import Membre
from django.contrib.staticfiles import finders

# Fonction pour convertir les nombres en lettres (version simplifiée)
def nombre_en_lettres(n):
    """
    Convertit un nombre en lettres en français avec une majuscule au début.
    """
    if not isinstance(n, (int, float)) or n < 0:
        return "Nombre invalide"
    
    # Pour les entiers uniquement (pas besoin de décimales pour l'historique)
    units = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
    teens = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
    tens = ["", "dix", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante-dix", "quatre-vingt", "quatre-vingt-dix"]

    if n == 0:
        return "Zéro"
    elif n < 10:
        return units[n].capitalize()
    elif 10 <= n < 20:
        return teens[n - 10].capitalize()
    elif 20 <= n < 100:
        if n % 10 == 0:
            return tens[n // 10].capitalize()
        elif n // 10 == 7 or n // 10 == 9:
            base = tens[n // 10 - 1] if n // 10 == 7 else tens[n // 10 - 1]
            return f"{base.capitalize()}-{teens[n % 10 - 10]}"
        else:
            separator = "-et-" if n % 10 == 1 and n // 10 != 8 else "-"
            return f"{tens[n // 10].capitalize()}{separator}{units[n % 10]}"
    else:
        return f"{n}"

# Fonction pour générer un QR code
def generer_qr_code(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    img.save(temp_file.name)
    return temp_file.name

def historique_membre_pdf(request, membre_id):
    membre = get_object_or_404(Membre, id=membre_id)

    # Préparer la réponse HTTP avec en-tête PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historique_{membre.nom}.pdf"'

    # Créer un buffer pour le PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Définition des marges
    MARGIN_LEFT = 1.5 * cm
    MARGIN_RIGHT = 1.5 * cm
    MARGIN_TOP = 1.5 * cm
    MARGIN_BOTTOM = 1.5 * cm

    # Couleurs de la charte graphique
    primary_color = colors.HexColor("#2c3e50")  # Bleu foncé
    secondary_color = colors.HexColor("#3498db")  # Bleu
    accent_color = colors.HexColor("#e74c3c")  # Rouge
    light_color = colors.HexColor("#ecf0f1")  # Gris très clair
    dark_color = colors.HexColor("#2c3e50")  # Texte foncé

    # Styles de texte
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=primary_color,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=dark_color,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        textColor=dark_color,
        fontName='Helvetica'
    )

    # En-tête avec deux colonnes (logo + infos société)
    p.setFillColor(light_color)
    p.rect(0, height - 3.5*cm, width, 3.5*cm, fill=True, stroke=False)
    
    # Logo (à gauche)
    try:
        logo_path = finders.find('images/logo1.png')
        if logo_path:  # si trouvé
            p.drawImage(logo_path, MARGIN_LEFT, height - 3.5*cm, width=2.5*cm, height=2.5*cm, mask='auto')
        else:
            raise FileNotFoundError("Logo non trouvé")
    except:
        # Placeholder
        p.setFillColor(secondary_color)
        p.roundRect(MARGIN_LEFT, height - 3*cm, 2.5*cm, 2.5*cm, 5, stroke=False, fill=True)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.2*cm, "LOGO")
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.6*cm, "RENEMICO")

    # Informations de la société (à droite)
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width - MARGIN_RIGHT, height - 2*cm, "REGROUPEMENT DES NEGOCIANTS MINIERS DU CONGO")
    
    p.setFont("Helvetica", 10)
    p.drawRightString(width - MARGIN_RIGHT, height - 2.7*cm, "297, Avenue Lubudi, Quartier Industriel, Manika, Kolwezi")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.1*cm, "Lualaba - RDC")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.5*cm, "Tél: +243 81 60 69 861 - Email: contact@renemico.com")

    # Position verticale initiale après l'en-tête
    y = height - 4.5*cm

    # Titre du document avec bandeau coloré
    p.setFillColor(secondary_color)
    p.rect(0, y - 0.5*cm, width, 1.2*cm, fill=True, stroke=False)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, y, "HISTORIQUE DES COTISATIONS")
    y -= 1.8*cm

    # Informations du membre
    p.setFillColor(primary_color)
    p.setFont("Helvetica-Bold", 12)

    titre = f"HISTORIQUE DES COTISATIONS DE: {membre.nom.upper()} {membre.post_nom.upper()} {membre.prenom.upper()}"

    # centrer horizontalement sur la largeur de la page
    p.drawCentredString(width / 2, y, titre)

    y -= 1*cm


    # Ligne de séparation
    p.setStrokeColor(secondary_color)
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    y -= 0.5*cm
    
    # Tableau des contributions
    data = [["Mois", "Montant (USD)"]]
    total = 0
    for c in membre.contributions.all().order_by('-mois'):
        data.append([c.mois.strftime('%B %Y'), f"{c.montant:,.2f}"])
        total += c.montant

    # Ajouter une ligne de total
    data.append(["TOTAL", f"{total:,.2f}"])

    # Largeur agrandie
    col_widths = [12*cm, 6*cm]  # élargir le tableau

    # Création du tableau
    table = Table(data, colWidths=col_widths)

    table.setStyle(TableStyle([
        # En-tête (centré, couleur)
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),  # centrer uniquement l’entête
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),

        # Colonne gauche (mois) → aligner à gauche sauf entête
        ('ALIGN', (0,1), (0,-1), 'LEFT'),

        # Colonne droite (montants) → aligner à droite sauf entête
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),

        # Ligne TOTAL (gras + couleur fond)
        ('BACKGROUND', (0,-1), (-1,-1), light_color),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.black),

        # Bordures du tableau
        ('GRID', (0,0), (-1,-2), 1, colors.HexColor("#bdc3c7")),
    ]))

    # Calculer la hauteur du tableau
    table_height = len(data) * 0.7 * cm
    table_width, _ = table.wrap(0, 0)

    # Position centrée horizontalement
    table_x = (width - table_width) / 2
    table_y = y - table_height

    # Vérifier si assez d’espace
    if table_y < MARGIN_BOTTOM + 5*cm:
        p.showPage()
        y = height - MARGIN_TOP
        table_y = y - table_height

    # Dessiner le tableau centré
    table.drawOn(p, table_x, table_y)
    y = table_y - 1*cm


    # Montant total en lettres
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(MARGIN_LEFT, y, "Total en lettres:")
    p.setFont("Helvetica", 10)
    
    montant_lettres = nombre_en_lettres(int(total)) + " dollars américains"
    p.drawString(MARGIN_LEFT + 3*cm, y, montant_lettres)
    y -= 1*cm

    # Générer et ajouter le QR Code
    qr_data = f"""
    RENEMICO - Historique des cotisations
    Membre: {membre.nom} {membre.post_nom} {membre.prenom}
    Total des cotisations: {total:.2f} USD
    Nombre de mois: {len(data)-2}
    Généré le: {timezone.now().strftime('%d/%m/%Y')}
    """
    
    qr_code_path = generer_qr_code(qr_data)
    p.drawImage(qr_code_path, MARGIN_LEFT, y - 3.5*cm, width=3.5*cm, height=3.5*cm)
    os.unlink(qr_code_path)  # Supprimer le fichier temporaire
    
    # Pied de page
    y = MARGIN_BOTTOM + 2.5*cm
    p.setStrokeColor(colors.HexColor('#bdc3c7'))
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    
    p.setFillColor(colors.HexColor('#7f8c8d'))
    p.setFont("Helvetica-Oblique", 9)
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.8*cm, "Merci pour votre confiance et vos cotisations !")
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.3*cm, "Ce document est généré automatiquement, pour toute question, contactez-nous à contact@renemico.com")

    p.showPage()
    p.save()
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


############################################################################################################
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

def historique_membre_excel(request, membre_id):
    membre = get_object_or_404(Membre, id=membre_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Historique {membre.nom}"

    # entêtes
    ws.append(["Mois", "Montant"])

    for c in membre.contributions.all().order_by('-mois'):
        ws.append([c.mois.strftime('%B %Y'), c.montant])

    # largeur colonnes
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    # sauvegarde dans mémoire
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # réponse HTTP
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="historique_{membre.nom}.xlsx"'
    return response

##########################################################################################################
#HISTORIQUE DES TOUS LES MEMBRES
# finance/views.py
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, FileResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import os
import qrcode
import tempfile
from django.conf import settings
from django.utils import timezone
from .models import Membre, Contribution
from django.contrib.staticfiles import finders

# Fonction pour convertir les nombres en lettres (version simplifiée)
def nombre_en_lettres(n):
    """
    Convertit un nombre en lettres en français avec une majuscule au début.
    """
    if not isinstance(n, (int, float)) or n < 0:
        return "Nombre invalide"
    
    # Pour les entiers uniquement (pas besoin de décimales pour l'historique)
    units = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf"]
    teens = ["dix", "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
    tens = ["", "dix", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante-dix", "quatre-vingt", "quatre-vingt-dix"]

    if n == 0:
        return "Zéro"
    elif n < 10:
        return units[n].capitalize()
    elif 10 <= n < 20:
        return teens[n - 10].capitalize()
    elif 20 <= n < 100:
        if n % 10 == 0:
            return tens[n // 10].capitalize()
        elif n // 10 == 7 or n // 10 == 9:
            base = tens[n // 10 - 1] if n // 10 == 7 else tens[n // 10 - 1]
            return f"{base.capitalize()}-{teens[n % 10 - 10]}"
        else:
            separator = "-et-" if n % 10 == 1 and n // 10 != 8 else "-"
            return f"{tens[n // 10].capitalize()}{separator}{units[n % 10]}"
    else:
        return f"{n}"

# Fonction pour générer un QR code
def generer_qr_code(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    img.save(temp_file.name)
    return temp_file.name

def tous_historique_pdf(request):
    # Préparer la réponse HTTP avec en-tête PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Historique_Tous_Membres.pdf"'

    # Créer un buffer pour le PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Définition des marges
    MARGIN_LEFT = 1.5 * cm
    MARGIN_RIGHT = 1.5 * cm
    MARGIN_TOP = 1.5 * cm
    MARGIN_BOTTOM = 1.5 * cm

    # Couleurs de la charte graphique
    primary_color = colors.HexColor("#2c3e50")  # Bleu foncé
    secondary_color = colors.HexColor("#3498db")  # Bleu
    accent_color = colors.HexColor("#e74c3c")  # Rouge
    light_color = colors.HexColor("#ecf0f1")  # Gris très clair
    dark_color = colors.HexColor("#2c3e50")  # Texte foncé

    # Styles de texte
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=primary_color,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=dark_color,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        textColor=dark_color,
        fontName='Helvetica'
    )

    # En-tête avec deux colonnes (logo + infos société)
    p.setFillColor(light_color)
    p.rect(0, height - 3.5*cm, width, 3.5*cm, fill=True, stroke=False)
    
    # Logo (à gauche)
    try:
        logo_path = finders.find('images/logo1.png')
        if logo_path:  # si trouvé
            p.drawImage(logo_path, MARGIN_LEFT, height - 3.5*cm, width=2.5*cm, height=2.5*cm, mask='auto')
        else:
            raise FileNotFoundError("Logo non trouvé")
    except:
        # Placeholder
        p.setFillColor(secondary_color)
        p.roundRect(MARGIN_LEFT, height - 3*cm, 2.5*cm, 2.5*cm, 5, stroke=False, fill=True)
        p.setFillColor(colors.white)
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.2*cm, "LOGO")
        p.drawCentredString(MARGIN_LEFT + 1.25*cm, height - 2.6*cm, "RENEMICO")

    # Informations de la société (à droite)
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width - MARGIN_RIGHT, height - 2*cm, "REGROUPEMENT DES NEGOCIANTS MINIERS DU CONGO")
    
    p.setFont("Helvetica", 10)
    p.drawRightString(width - MARGIN_RIGHT, height - 2.7*cm, "297, Avenue Lubudi, Quartier Industriel, Manika, Kolwezi")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.1*cm, "Lualaba - RDC")
    p.drawRightString(width - MARGIN_RIGHT, height - 3.5*cm, "Tél: +243 81 60 69 861 - Email: contact@renemico.com")

    # Position verticale initiale après l'en-tête
    y = height - 4.5*cm

    # Titre du document avec bandeau coloré
    p.setFillColor(secondary_color)
    p.rect(0, y - 0.5*cm, width, 1.2*cm, fill=True, stroke=False)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width/2, y, "HISTORIQUE DES COTISATIONS")
    y -= 1.8*cm

    # Informations du document
    p.setFillColor(primary_color)
    p.setFont("Helvetica-Bold", 12)

    titre = "HISTORIQUE DES COTISATIONS DE TOUS LES MEMBRES"

    # centrer horizontalement sur la largeur de la page
    p.drawCentredString(width / 2, y, titre)

    y -= 1*cm

    # Ligne de séparation
    p.setStrokeColor(secondary_color)
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    y -= 0.5*cm
    
    # Récupérer toutes les contributions (avec membre)
    contributions = Contribution.objects.select_related('membre').order_by('membre__nom', '-mois')
    
    # Tableau des contributions
    data = [["Membre", "Code", "Province", "Mois", "Montant (USD)"]]
    total_general = 0
    
    for c in contributions:
        data.append([
            f"{c.membre.nom} {c.membre.post_nom}",
            c.membre.code,
            c.membre.province,
            c.mois.strftime('%B %Y'),
            f"{c.montant:,.2f}"
        ])
        total_general += c.montant

    # Ajouter une ligne de total
    data.append([
        "TOTAL GENERAL", "", "", "",  # occupe les 4 premières colonnes
        f"{total_general:,.2f}"       # montant dans la dernière colonne
    ])

    # Largeur des colonnes
    col_widths = [4*cm, 3*cm, 4*cm, 5*cm, 3*cm]

    # Création du tableau
    table = Table(data, colWidths=col_widths)

    table.setStyle(TableStyle([
        # En-tête
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),

        # Colonnes (alignement normal)
        ('ALIGN', (0,1), (0,-2), 'LEFT'),     # Membre
        ('ALIGN', (1,1), (1,-2), 'CENTER'),   # Code
        ('ALIGN', (2,1), (2,-2), 'LEFT'),     # Province
        ('ALIGN', (3,1), (3,-2), 'LEFT'),     # Mois
        ('ALIGN', (4,1), (4,-2), 'RIGHT'),    # Montant

        # Ligne TOTAL (fusion + style)
        ('SPAN', (0,-1), (3,-1)),             # fusionner colonnes 0→3
        ('ALIGN', (0,-1), (0,-1), 'LEFT'),    # "TOTAL GENERAL" aligné à gauche
        ('ALIGN', (4,-1), (4,-1), 'RIGHT'),   # montant aligné à droite
        ('BACKGROUND', (0,-1), (-1,-1), light_color),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LINEABOVE', (0,-1), (-1,-1), 1, colors.black),

        # Bordures
        ('GRID', (0,0), (-1,-2), 1, colors.HexColor("#bdc3c7")),
        ('FONTSIZE', (0,1), (-1,-1), 9),
    ]))
    # Calculer la hauteur du tableau
    table_height = len(data) * 0.7 * cm
    table_width, _ = table.wrap(0, 0)

    # Position centrée horizontalement
    table_x = (width - table_width) / 2
    table_y = y - table_height

    # Vérifier si assez d'espace
    if table_y < MARGIN_BOTTOM + 5*cm:
        p.showPage()
        y = height - MARGIN_TOP
        table_y = y - table_height

    # Dessiner le tableau centré
    table.drawOn(p, table_x, table_y)
    y = table_y - 1*cm

    # Montant total en lettres
    p.setFillColor(dark_color)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(MARGIN_LEFT, y, "Total général en lettres:")
    p.setFont("Helvetica", 10)
    
    montant_lettres = nombre_en_lettres(int(total_general)) + " dollars américains"
    p.drawString(MARGIN_LEFT + 4*cm, y, montant_lettres)
    y -= 1*cm

    # Générer et ajouter le QR Code
    qr_data = f"""
    RENEMICO - Historique des cotisations de tous les membres
    Total des cotisations: {total_general:.2f} USD
    Nombre de contributions: {len(data)-2}
    Généré le: {timezone.now().strftime('%d/%m/%Y')}
    """
    
    qr_code_path = generer_qr_code(qr_data)
    p.drawImage(qr_code_path, MARGIN_LEFT, y - 3.5*cm, width=3.5*cm, height=3.5*cm)
    os.unlink(qr_code_path)  # Supprimer le fichier temporaire
    
    # Pied de page
    y = MARGIN_BOTTOM + 2.5*cm
    p.setStrokeColor(colors.HexColor('#bdc3c7'))
    p.setLineWidth(1)
    p.line(MARGIN_LEFT, y, width - MARGIN_RIGHT, y)
    
    p.setFillColor(colors.HexColor('#7f8c8d'))
    p.setFont("Helvetica-Oblique", 9)
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.8*cm, "Merci pour votre confiance et vos cotisations !")
    p.drawCentredString(width/2, MARGIN_BOTTOM + 1.3*cm, "Ce document est généré automatiquement, pour toute question, contactez-nous à contact@renemico.com")

    p.showPage()
    p.save()
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

###############################################################################################
import openpyxl
from openpyxl.utils import get_column_letter

def tous_historique_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique Contributions"

    # Entêtes
    ws.append(["Membre", "Code", "Mois", "Montant"])

    contributions = Contribution.objects.select_related('membre').order_by('membre__nom', '-mois')

    for c in contributions:
        ws.append([
            f"{c.membre.nom} {c.membre.post_nom}",
            c.membre.code,
            c.mois.strftime('%B %Y'),
            c.montant
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Historique_Tous_Membres.xlsx")


######################################################################################################
from django.shortcuts import render, redirect
from django.db.models import Sum
from .models import Operation
from .forms import OperationForm

def liste_operations(request):
    operations = Operation.objects.all()
    total_entrees = Operation.objects.filter(type_operation='ENTREE').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sorties = Operation.objects.filter(type_operation='SORTIE').aggregate(Sum('montant'))['montant__sum'] or 0
    solde = total_entrees - total_sorties
    return render(request, "liste_operations.html", {
        "operations": operations,
        "total_entrees": total_entrees,
        "total_sorties": total_sorties,
        "solde": solde
    })
###############################################################################################################

def ajouter_operation(request):
    if request.method == "POST":
        form = OperationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("finance:liste_operations")
    else:
        form = OperationForm()
    return render(request, "ajouter_operation.html", {"form": form})

################################################################################################################
def modifier_operation(request, pk):
    operation = get_object_or_404(Operation, pk=pk)
    if request.method == "POST":
        form = OperationForm(request.POST, instance=operation)
        if form.is_valid():
            form.save()
            return redirect("finance:liste_operations")
    else:
        form = OperationForm(instance=operation)
    return render(request, "modifier_operation.html", {"form": form, "operation": operation})

##########################################################################################################
def supprimer_operation(request, pk):
    operation = get_object_or_404(Operation, pk=pk)
    if request.method == "POST":
        operation.delete()
        return redirect("finance:liste_operations")
    return render(request, "supprimer_operation.html", {"operation": operation})

def detail_operation(request, pk):
    operation = get_object_or_404(Operation, pk=pk)
    return render(request, "detail_operation.html", {"operation": operation})

#######################################################################################################
import datetime
from django.http import HttpResponse
from django.db.models import Sum
from .models import Operation
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

# --- Vue pour exporter en Excel ---
def export_excel_operations(request):
    operations = Operation.objects.all()

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opérations"

    # En-tête
    ws.append(["Date", "Type", "Description", "Montant (USD)"])

    # Contenu
    for op in operations:
        ws.append([op.date.strftime("%d/%m/%Y"), op.type_operation, op.motif, op.montant])


    # Réponse HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=operations.xlsx'
    wb.save(response)
    return response

################################################################################################################
# --- Vue pour exporter en PDF ---
def export_pdf_operations(request, periode):
    today = datetime.date.today()
    queryset = Operation.objects.all()

    # Filtrer selon la période demandée
    if periode == "journalier":
        queryset = queryset.filter(date=today)
    elif periode == "mensuel":
        queryset = queryset.filter(date__year=today.year, date__month=today.month)
    elif periode == "trimestriel":
        trimestre = (today.month - 1) // 3 + 1
        mois_debut = 3 * (trimestre - 1) + 1
        mois_fin = mois_debut + 2
        queryset = queryset.filter(
            date__year=today.year,
            date__month__gte=mois_debut,
            date__month__lte=mois_fin
        )
    elif periode == "semestriel":
        semestre = 1 if today.month <= 6 else 2
        mois_debut = 1 if semestre == 1 else 7
        mois_fin = 6 if semestre == 1 else 12
        queryset = queryset.filter(
            date__year=today.year,
            date__month__gte=mois_debut,
            date__month__lte=mois_fin
        )
    elif periode == "annuel":
        queryset = queryset.filter(date__year=today.year)

    # Calcul des totaux
    total_entrees = queryset.filter(type_operation='ENTREE').aggregate(Sum('montant'))['montant__sum'] or 0
    total_sorties = queryset.filter(type_operation='SORTIE').aggregate(Sum('montant'))['montant__sum'] or 0
    solde = total_entrees - total_sorties

    # Création du PDF
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename=rapport_{periode}.pdf'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # En-tête du document
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height-2*cm, f"Rapport {periode.upper()} des Opérations")

    # En-têtes de colonnes
    p.setFont("Helvetica", 10)
    y = height - 3*cm
    p.drawString(2*cm, y, "Date")
    p.drawString(5*cm, y, "Type")
    p.drawString(9*cm, y, "Description")
    p.drawString(15*cm, y, "Montant")
    y -= 0.5*cm

    # Données des opérations
    for op in queryset:
        p.drawString(2*cm, y, op.date.strftime("%d/%m/%Y"))
        p.drawString(5*cm, y, op.type_operation)
        p.drawString(9*cm, y, op.motif[:30])  # Tronquer si trop long
        p.drawRightString(19*cm, y, f"{op.montant:,.2f}")
        y -= 0.5*cm
        
        # Nouvelle page si nécessaire
        if y < 3*cm:
            p.showPage()
            y = height - 2*cm

    # Section des totaux
    y -= 1*cm
    p.setFont("Helvetica-Bold", 10)
    p.drawString(2*cm, y, f"TOTAL ENTRÉES: {total_entrees:,.2f} USD")
    y -= 0.5*cm
    p.drawString(2*cm, y, f"TOTAL SORTIES: {total_sorties:,.2f} USD")
    y -= 0.5*cm
    p.drawString(2*cm, y, f"SOLDE: {solde:,.2f} USD")

    p.save()
    return response
